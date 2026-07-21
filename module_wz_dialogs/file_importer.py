import contextlib
import glob
import os
import csv
import shutil
import subprocess
import tempfile

from qgis.PyQt.QtCore import Qt, QVariant
from qgis.PyQt.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QComboBox, QGroupBox, QGridLayout,
    QMessageBox, QWidget, QStackedWidget, QLineEdit,
)
from qgis.core import (
    QgsProject, QgsVectorLayer, QgsRasterLayer,
    QgsField, QgsFeature, QgsGeometry, QgsPointXY,
    QgsCoordinateReferenceSystem,
)
from qgis.gui import QgsProjectionSelectionWidget

from ..modules.layer_utils import add_to_plugin_group


class FileImportDialog(QDialog):
    """Import CSV, XLSX, DWG/DXF, or PDF files as QGIS layers."""

    _SUPPORTED = {
        '.csv':  'csv',
        '.xlsx': 'xlsx',
        '.xls':  'xlsx',
        '.dwg':  'dwg',
        '.dxf':  'dwg',
        '.pdf':  'pdf',
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Import File')
        self.setMinimumWidth(500)
        self._file_path = ''
        self._fmt = ''
        self._build_ui()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── File picker ───────────────────────────────────────────────
        file_group = QGroupBox('File')
        file_row = QHBoxLayout(file_group)
        self._path_label = QLabel('No file selected')
        self._path_label.setWordWrap(True)
        browse_btn = QPushButton('Browse…')
        browse_btn.setFixedWidth(80)
        browse_btn.clicked.connect(self._browse)
        file_row.addWidget(self._path_label, 1)
        file_row.addWidget(browse_btn)
        layout.addWidget(file_group)

        # ── Format-specific options ───────────────────────────────────
        self._options_group = QGroupBox('Options')
        self._stack = QStackedWidget()
        opt_layout = QVBoxLayout(self._options_group)
        opt_layout.addWidget(self._stack)
        layout.addWidget(self._options_group)

        # Page 0 — placeholder (no file yet)
        self._stack.addWidget(QLabel('Select a file above.'))

        # Page 1 — CSV / XLSX
        csv_widget = QWidget()
        csv_grid = QGridLayout(csv_widget)
        csv_grid.setColumnStretch(1, 1)
        csv_grid.addWidget(QLabel('X column:'), 0, 0)
        self._x_combo = QComboBox()
        csv_grid.addWidget(self._x_combo, 0, 1)
        csv_grid.addWidget(QLabel('Y column:'), 1, 0)
        self._y_combo = QComboBox()
        csv_grid.addWidget(self._y_combo, 1, 1)
        csv_grid.addWidget(QLabel('CRS:'), 2, 0)
        self._csv_crs = QgsProjectionSelectionWidget()
        self._csv_crs.setCrs(QgsCoordinateReferenceSystem('EPSG:4326'))
        csv_grid.addWidget(self._csv_crs, 2, 1)
        self._stack.addWidget(csv_widget)

        # Page 2 — DWG / DXF
        dwg_widget = QWidget()
        dwg_layout = QVBoxLayout(dwg_widget)
        dwg_layout.setContentsMargins(0, 0, 0, 0)
        dwg_layout.addWidget(QLabel(
            'DWG/DXF will be imported as a vector layer via OGR.\n'
            'DXF is supported on all platforms; DWG requires the ODA/OpenDWG library.'
        ))
        dwg_layout.addStretch()
        self._stack.addWidget(dwg_widget)

        # Page 3 — PDF
        pdf_widget = QWidget()
        pdf_layout = QVBoxLayout(pdf_widget)
        pdf_layout.setContentsMargins(0, 0, 0, 0)
        pdf_layout.addWidget(QLabel(
            'PDF will be imported as a raster layer via GDAL.\n'
            'Georeferenced PDFs retain their spatial extent automatically.'
        ))
        pdf_layout.addStretch()
        self._stack.addWidget(pdf_widget)

        # ── Layer name ────────────────────────────────────────────────
        name_row = QHBoxLayout()
        name_row.addWidget(QLabel('Layer name:'))
        self._name_edit = QLineEdit()
        name_row.addWidget(self._name_edit)
        layout.addLayout(name_row)

        # ── Action buttons ────────────────────────────────────────────
        btn_row = QHBoxLayout()
        self._import_btn = QPushButton('Import')
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        close_btn = QPushButton('Close')
        close_btn.clicked.connect(self.close)
        btn_row.addStretch()
        btn_row.addWidget(self._import_btn)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

    # ------------------------------------------------------------------
    # File browser
    # ------------------------------------------------------------------

    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, 'Open File', '',
            'Supported files (*.csv *.xlsx *.xls *.pdf);; '
            'CSV (*.csv);;Excel (*.xlsx *.xls);;PDF (*.pdf)'
        )
        if not path:
            return

        self._file_path = path
        self._path_label.setText(path)
        ext = os.path.splitext(path)[1].lower()
        self._fmt = self._SUPPORTED.get(ext, '')
        base = os.path.splitext(os.path.basename(path))[0]
        self._name_edit.setText(base)

        if self._fmt == 'csv':
            self._load_csv_headers(path)
            self._stack.setCurrentIndex(1)
            self._import_btn.setEnabled(True)
        elif self._fmt == 'xlsx':
            self._load_xlsx_headers(path)
            self._stack.setCurrentIndex(1)
            self._import_btn.setEnabled(True)
        elif self._fmt == 'dwg':
            self._stack.setCurrentIndex(2)
            self._import_btn.setEnabled(True)
        elif self._fmt == 'pdf':
            self._stack.setCurrentIndex(3)
            self._import_btn.setEnabled(True)
        else:
            self._stack.setCurrentIndex(0)
            self._import_btn.setEnabled(False)
            QMessageBox.warning(self, 'Unsupported', f'File type "{ext}" is not supported.')

    # ------------------------------------------------------------------
    # Header detection
    # ------------------------------------------------------------------

    def _load_csv_headers(self, path):
        self._x_combo.clear()
        self._y_combo.clear()
        try:
            with open(path, newline='', encoding='utf-8-sig') as f:
                headers = next(csv.reader(f))
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Could not read CSV headers:\n{e}')
            return
        self._x_combo.addItems(headers)
        self._y_combo.addItems(headers)
        self._auto_pick_xy(headers)

    def _load_xlsx_headers(self, path):
        self._x_combo.clear()
        self._y_combo.clear()
        headers = self._xlsx_headers(path)
        if not headers:
            return
        self._x_combo.addItems(headers)
        self._y_combo.addItems(headers)
        self._auto_pick_xy(headers)

    def _auto_pick_xy(self, headers):
        x_names = {'x', 'easting', 'lon', 'longitude', 'east', 'e'}
        y_names = {'y', 'northing', 'lat', 'latitude', 'north', 'n'}
        for i, h in enumerate(headers):
            hl = h.lower().strip()
            if hl in x_names:
                self._x_combo.setCurrentIndex(i)
            if hl in y_names:
                self._y_combo.setCurrentIndex(i)

    @staticmethod
    def _xlsx_headers(path):
        """Return column headers from an XLSX file; try openpyxl then OGR."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [
                str(cell.value) if cell.value is not None else f'Col{i}'
                for i, cell in enumerate(next(ws.iter_rows()))
            ]
            wb.close()
            return headers
        except ImportError:
            pass
        with contextlib.suppress(Exception):
            from osgeo import ogr
            ds = ogr.Open(path)
            if ds:
                lyr = ds.GetLayer(0)
                if lyr:
                    defn = lyr.GetLayerDefn()
                    return [defn.GetFieldDefn(i).GetName() for i in range(defn.GetFieldCount())]
        QMessageBox.critical(
            None, 'Error',
            'Could not read XLSX headers.\n'
            'Install openpyxl (pip install openpyxl) or ensure GDAL/OGR has XLSX support.'
        )
        return []

    # ------------------------------------------------------------------
    # Import dispatch
    # ------------------------------------------------------------------

    def _do_import(self):
        name = self._name_edit.text().strip() or 'imported'
        if self._fmt in ('csv', 'xlsx'):
            self._import_tabular(name)
        elif self._fmt == 'dwg':
            self._import_dwg(name)
        elif self._fmt == 'pdf':
            self._import_pdf(name)

    # ------------------------------------------------------------------
    # CSV / XLSX → point layer
    # ------------------------------------------------------------------

    def _import_tabular(self, name):
        x_col = self._x_combo.currentText()
        y_col = self._y_combo.currentText()
        crs   = self._csv_crs.crs()

        if self._fmt == 'csv':
            rows = self._read_csv_rows()
        else:
            rows = self._read_xlsx_rows()
        if rows is None:
            return

        extra = [k for k in (rows[0].keys() if rows else []) if k not in (x_col, y_col)]

        fields_part = '&'.join(f'field={f}:string' for f in extra)
        uri = f'Point?crs={crs.authid()}'
        if fields_part:
            uri += f'&{fields_part}'

        mem_lyr = QgsVectorLayer(uri, name, 'memory')
        if not mem_lyr.isValid():
            QMessageBox.critical(self, 'Error', 'Could not create memory layer.')
            return

        mem_lyr.startEditing()
        ok = 0
        for row in rows:
            try:
                x = float(row[x_col])
                y = float(row[y_col])
            except (ValueError, KeyError, TypeError):
                continue
            feat = QgsFeature(mem_lyr.fields())
            feat.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(x, y)))
            for f in extra:
                idx = mem_lyr.fields().indexOf(f)
                if idx >= 0:
                    feat.setAttribute(idx, str(row.get(f, '') or ''))
            mem_lyr.addFeature(feat)
            ok += 1
        mem_lyr.commitChanges()

        if ok == 0:
            QMessageBox.warning(self, 'No data',
                f'No valid rows found — check that "{x_col}" and "{y_col}" contain numbers.')
            return

        add_to_plugin_group(mem_lyr)
        QMessageBox.information(self, 'Done', f'Imported {ok} points as "{name}".')

    def _read_csv_rows(self):
        try:
            with open(self._file_path, newline='', encoding='utf-8-sig') as f:
                return list(csv.DictReader(f))
        except Exception as e:
            QMessageBox.critical(self, 'Error', str(e))
            return None

    def _read_xlsx_rows(self):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [
                str(v) if v is not None else f'Col{i}'
                for i, v in enumerate(next(rows_iter))
            ]
            result = [
                {headers[i]: (str(v) if v is not None else '') for i, v in enumerate(row)}
                for row in rows_iter
            ]
            wb.close()
            return result
        except ImportError:
            pass
        # OGR fallback
        try:
            from osgeo import ogr
            ds = ogr.Open(self._file_path)
            if ds:
                lyr = ds.GetLayer(0)
                if lyr:
                    defn = lyr.GetLayerDefn()
                    names = [defn.GetFieldDefn(i).GetName() for i in range(defn.GetFieldCount())]
                    return [
                        {n: str(feat.GetField(n) or '') for n in names}
                        for feat in lyr
                    ]
        except Exception as e:
            QMessageBox.critical(self, 'Error', str(e))
            return None
        QMessageBox.critical(self, 'Error', 'Cannot read XLSX — install openpyxl.')
        return None

    # ------------------------------------------------------------------
    # DWG / DXF → vector layer
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # ODA File Converter helpers (DWG → DXF)
    # ------------------------------------------------------------------

    @staticmethod
    def _find_oda_converter():
        """Return the path to ODAFileConverter.exe, or None if not installed."""
        patterns = [
            r'C:\Program Files\ODA\ODAFileConverter*\ODAFileConverter.exe',
            r'C:\Program Files (x86)\ODA\ODAFileConverter*\ODAFileConverter.exe',
            r'C:\Program Files\Teigha\TeighaFileConverter*\TeighaFileConverter.exe',
            r'C:\Program Files (x86)\Teigha\TeighaFileConverter*\TeighaFileConverter.exe',
        ]
        for pat in patterns:
            hits = sorted(glob.glob(pat))
            if hits:
                return hits[-1]  # newest version
        return shutil.which('ODAFileConverter') or shutil.which('TeighaFileConverter')

    def _dwg_to_dxf(self, dwg_path):
        """Convert a DWG file to DXF beside the original; return the DXF path or None."""
        oda = self._find_oda_converter()
        if not oda:
            return None

        tmp_in  = tempfile.mkdtemp(prefix='ugsurv_in_')
        tmp_out = tempfile.mkdtemp(prefix='ugsurv_out_')
        try:
            shutil.copy2(dwg_path, os.path.join(tmp_in, os.path.basename(dwg_path)))
            subprocess.run(
                [oda, tmp_in, tmp_out, 'ACAD2018', 'DXF', '0', '1'],
                timeout=60, check=False,
            )
            base    = os.path.splitext(os.path.basename(dwg_path))[0]
            dxf_out = os.path.join(tmp_out, base + '.dxf')
            if os.path.exists(dxf_out):
                dest = os.path.join(os.path.dirname(dwg_path), base + '.dxf')
                shutil.copy2(dxf_out, dest)
                return dest
        except Exception:
            pass
        finally:
            shutil.rmtree(tmp_in,  ignore_errors=True)
            shutil.rmtree(tmp_out, ignore_errors=True)
        return None

    def _dwg_fallback_dialog(self, dwg_path):
        """Show an error/options dialog when DWG auto-conversion fails.

        Returns a DXF path the user browses to, or None to abort.
        """
        oda_found = self._find_oda_converter() is not None

        msg = QMessageBox(self)
        msg.setWindowTitle('Cannot open DWG')
        msg.setIcon(QMessageBox.Warning)

        if oda_found:
            msg.setText(
                'ODA File Converter was found but the conversion failed.\n'
                'The DWG file may be corrupt or use an unsupported version.\n\n'
                'Convert the file manually to DXF, then click "Browse for DXF".'
            )
        else:
            msg.setText(
                'DWG files need the free ODA File Converter to open automatically.\n\n'
                'Option A — Install ODA File Converter, then retry:\n'
                'opendesign.com/guestfiles/oda_file_converter\n\n'
                'Option B — Convert the DWG to DXF using any CAD tool\n'
                '(FreeCAD, LibreCAD, AutoCAD, online converter, etc.),\n'
                'then click "Browse for DXF" below.'
            )

        browse_btn = msg.addButton('Browse for DXF…', QMessageBox.AcceptRole)
        msg.addButton(QMessageBox.Cancel)
        msg.exec_()

        if msg.clickedButton() != browse_btn:
            return None

        dxf_path, _ = QFileDialog.getOpenFileName(
            self, 'Select the DXF version of the drawing',
            os.path.dirname(dwg_path),
            'DXF files (*.dxf)',
        )
        return dxf_path or None

    # ------------------------------------------------------------------
    # DWG / DXF → vector layer
    # ------------------------------------------------------------------

    def _import_dwg(self, name):
        path = self._file_path
        ext  = os.path.splitext(path)[1].lower()

        # 1. Try opening directly — works when GDAL includes libopencad (open-source
        #    DWG reader built into many QGIS/OSGeo4W distributions).
        probe = QgsVectorLayer(path, name, 'ogr')

        if not probe.isValid() and ext == '.dwg':
            # 2. libopencad unavailable or version unsupported — try ODA File Converter.
            dxf = self._dwg_to_dxf(path)
            if dxf:
                path  = dxf
                probe = QgsVectorLayer(path, name, 'ogr')
            else:
                # 3. Neither worked — let user browse for a pre-converted DXF.
                path = self._dwg_fallback_dialog(path)
                if not path:
                    return
                probe = QgsVectorLayer(path, name, 'ogr')

        if not probe.isValid():
            QMessageBox.critical(self, 'Error',
                f'Could not open "{os.path.basename(path)}".')
            return

        # DXF files expose multiple OGR sublayers — load each non-empty one.
        SEP = '!!::!!'
        sublayers = probe.dataProvider().subLayers()
        if not sublayers:
            add_to_plugin_group(probe)
            QMessageBox.information(self, 'Done', f'Loaded "{name}".')
            return

        loaded = 0
        for sl in sublayers:
            parts = sl.split(SEP)
            if len(parts) < 3:
                continue
            sub_name, feat_count = parts[1], parts[2]
            if feat_count.isdigit() and int(feat_count) == 0:
                continue
            sub_lyr = QgsVectorLayer(f'{path}|layername={sub_name}',
                                     f'{name} — {sub_name}', 'ogr')
            if sub_lyr.isValid():
                add_to_plugin_group(sub_lyr)
                loaded += 1

        if loaded == 0:
            add_to_plugin_group(probe)
            loaded = 1

        QMessageBox.information(self, 'Done',
            f'Loaded {loaded} layer(s) from "{os.path.basename(path)}".')

    # ------------------------------------------------------------------
    # PDF → raster layer
    # ------------------------------------------------------------------

    def _import_pdf(self, name):
        lyr = QgsRasterLayer(self._file_path, name)
        if not lyr.isValid():
            QMessageBox.critical(self, 'Error',
                f'Could not load PDF as raster.\n{lyr.error().message()}')
            return
        add_to_plugin_group(lyr)
        QMessageBox.information(self, 'Done', f'Loaded "{name}" as raster layer.')
